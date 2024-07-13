from selenium import webdriver
from bs4 import BeautifulSoup
import time
import openpyxl
from openpyxl import load_workbook
import os

#this function is used to check the file and the current data is already present or not 
def append_to_excel(filename, data):
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "ONEPLUS mobile sales Data"
        worksheet.append(['ONEPLUS MOBILE NAME ', 'MOBILE RATING', 'MOBILE  RATING COUNT', 'MOBILE  REVIEW COUNT', 'MOBILE RATE','PRODUCT LINK'])
    
    worksheet.append(data)
    workbook.save(filename)

workbook=openpyxl.Workbook()
worksheet=workbook.active

worksheet.append(['ONEPLUS MOBILE NAME ', 'MOBILE RATING', 'MOBILE  RATING COUNT', 'MOBILE  REVIEW COUNT', 'MOBILE RATE','PRODUCT LINK'])

#scraping website url
url = 'https://www.flipkart.com/mobiles/pr?sid=tyy%2C4io&param=167811&ctx=eyJjYXJkQ29udGV4dCI6eyJhdHRyaWJ1dGVzIjp7InRpdGxlIjp7Im11bHRpVmFsdWVkQXR0cmlidXRlIjp7ImtleSI6InRpdGxlIiwiaW5mZXJlbmNlVHlwZSI6IlRJVExFIiwidmFsdWVzIjpbIkFwcGxlIFNtYXJ0cGhvbmVzIl0sInZhbHVlVHlwZSI6Ik1VTFRJX1ZBTFVFRCJ9fX19fQ%3D%3D&wid=27.productCard.PMU_V2_24&sort=popularity&p%5B%5D=facets.brand%255B%255D%3DOnePlus'
        
#Path to the ChromeDriver executable
path = r'C:/Program Files (x86)/chromedriver'

#Initialize the WebDriver
browser = webdriver.Chrome(executable_path=path)
browser.get(url)

#Wait for the page to fully load
time.sleep(20)

#Parse the page source with BeautifulSoup
soup = BeautifulSoup(browser.page_source, 'html.parser')

#Close the WebDriver
browser.close()

# Find all tables with the specified class
tables=soup.find_all('div', class_='_75nlfW')
for table in tables:
    product_name=table.find('div',class_='KzDlHZ').text
    product_ratings=table.find('div', class_='XQDdHH').text
    if product_ratings:
        product_rating=product_ratings
    else:
        # If the rating is not found, print a message or handle it as needed
        product_rating="Product rate not found"
        continue 
    product_reviews=table.find('span',class_='Wphh3N').text
    product_review=product_reviews.split('\xa0&\xa0')
    product_rating_value=product_review[0]
    product_review_value=product_review[1]
    product_rates=table.find('div',class_='Nx9bqj _4b5DiR').text
    if product_rates:
        
        product_rate=product_rates
    else:
        # If the rating is not found, print a message or handle it as needed
        product_rate="Product rate not found"
        continue
    product_links=table.find('a',class_='CGtC98')
    product_href=product_links.get('href')
    product_link='https://www.flipkart.com'+product_href

# Create a list to hold all elements in sequence
    mobile_data=[product_name, product_rating, product_rating_value, product_review_value,product_rate,product_link]

    filename = "ONEPLUS mobile sales Data.xlsx"
    append_to_excel(filename, mobile_data)
